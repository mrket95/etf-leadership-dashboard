#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ETF Leadership Dashboard

Purpose:
- Download daily OHLC data for ETF tickers.
- Build an interactive HTML dashboard where traces can be toggled on/off.
- User can enter a period, choose benchmark, and see period high/low/worst drawdown metrics.
- Period high uses max daily High; period low uses min daily Low.
- Worst drawdown uses daily Low divided by prior running daily High.

Outputs:
- output/prices_ohlc.csv
- output/download_status.csv
- output/default_period_summary.csv
- output/dashboard.html
- etf_leadership_dashboard.xlsx
"""

from __future__ import annotations

import argparse
import json
import math
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import urllib.parse

import pandas as pd
import requests


ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "output"
CACHE = ROOT / ".cache_prices"
OUTPUT.mkdir(exist_ok=True)
CACHE.mkdir(exist_ok=True)

LOG_PATH = OUTPUT / "run_log.txt"


def log(msg: str) -> None:
    line = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(line, flush=True)
    with LOG_PATH.open("a", encoding="utf-8") as f:
        f.write(line + "\n")


def load_config(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def valid_ohlc(df: pd.DataFrame) -> bool:
    if df is None or df.empty:
        return False
    cols = {c.lower() for c in df.columns}
    return {"date", "open", "high", "low", "close"}.issubset(cols)


def normalize_ohlc(df: pd.DataFrame, ticker: str, source: str) -> pd.DataFrame:
    # Accept typical Date/Open/High/Low/Close/Volume or lower-case variants
    rename = {}
    for c in df.columns:
        cl = c.lower()
        if cl == "date":
            rename[c] = "Date"
        elif cl == "open":
            rename[c] = "Open"
        elif cl == "high":
            rename[c] = "High"
        elif cl == "low":
            rename[c] = "Low"
        elif cl == "close":
            rename[c] = "Close"
        elif cl == "volume":
            rename[c] = "Volume"
    df = df.rename(columns=rename)
    keep = [c for c in ["Date", "Open", "High", "Low", "Close", "Volume"] if c in df.columns]
    df = df[keep].copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    for c in ["Open", "High", "Low", "Close", "Volume"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    if "Volume" not in df.columns:
        df["Volume"] = None
    df = df.dropna(subset=["Date", "Open", "High", "Low", "Close"]).sort_values("Date")
    df["Ticker"] = ticker
    df["Source"] = source
    return df[["Date", "Ticker", "Open", "High", "Low", "Close", "Volume", "Source"]]


def download_stooq(ticker: str, start_date: str, end_date: str, timeout: int = 20) -> pd.DataFrame:
    # Stooq US tickers use ticker.us
    s = f"{ticker.lower()}.us"
    d1 = start_date.replace("-", "")
    d2 = end_date.replace("-", "")
    url = f"https://stooq.com/q/d/l/?s={urllib.parse.quote(s)}&d1={d1}&d2={d2}&i=d"
    headers = {"User-Agent": "Mozilla/5.0 etf-leadership-dashboard"}
    r = requests.get(url, headers=headers, timeout=timeout)
    r.raise_for_status()
    text = r.text.strip()
    if not text or text.lower().startswith("no data"):
        raise ValueError("Stooq returned no data")
    from io import StringIO
    df = pd.read_csv(StringIO(text))
    df = normalize_ohlc(df, ticker, "stooq")
    if df.empty:
        raise ValueError("Stooq parsed empty data")
    return df


def to_unix(date_str: str) -> int:
    return int(pd.Timestamp(date_str, tz="UTC").timestamp())


def download_yahoo(ticker: str, start_date: str, end_date: str, timeout: int = 20) -> pd.DataFrame:
    p1 = to_unix(start_date)
    p2 = to_unix((pd.Timestamp(end_date) + pd.Timedelta(days=1)).strftime("%Y-%m-%d"))
    url = (
        f"https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(ticker)}"
        f"?period1={p1}&period2={p2}&interval=1d&events=history&includeAdjustedClose=true"
    )
    headers = {"User-Agent": "Mozilla/5.0 etf-leadership-dashboard"}
    r = requests.get(url, headers=headers, timeout=timeout)
    r.raise_for_status()
    data = r.json()
    result = data.get("chart", {}).get("result", [])
    if not result:
        raise ValueError("Yahoo returned no chart result")
    res = result[0]
    timestamps = res.get("timestamp", [])
    quote = res.get("indicators", {}).get("quote", [{}])[0]
    if not timestamps or not quote:
        raise ValueError("Yahoo missing timestamp/quote")
    df = pd.DataFrame({
        "Date": pd.to_datetime(timestamps, unit="s", utc=True).tz_convert(None).date,
        "Open": quote.get("open", []),
        "High": quote.get("high", []),
        "Low": quote.get("low", []),
        "Close": quote.get("close", []),
        "Volume": quote.get("volume", []),
    })
    df = normalize_ohlc(df, ticker, "yahoo")
    if df.empty:
        raise ValueError("Yahoo parsed empty data")
    return df


def get_ohlc_for_ticker(ticker: str, start_date: str, end_date: str, source_order: List[str]) -> Tuple[Optional[pd.DataFrame], dict]:
    status = {
        "Ticker": ticker,
        "Status": "",
        "SourceUsed": "",
        "Rows": 0,
        "FirstDate": "",
        "LastDate": "",
        "Message": "",
    }
    cache_path = CACHE / f"{ticker}.csv"
    last_msg = ""

    for source in source_order:
        try:
            log(f"Downloading {ticker} from {source}")
            if source == "stooq":
                df = download_stooq(ticker, start_date, end_date)
            elif source == "yahoo":
                df = download_yahoo(ticker, start_date, end_date)
            else:
                raise ValueError(f"Unknown source {source}")
            if valid_ohlc(df):
                df.to_csv(cache_path, index=False)
                log(f"{ticker} downloaded successfully from {source}: {len(df)} rows")
                status.update(
                    Status="OK",
                    SourceUsed=source,
                    Rows=len(df),
                    FirstDate=df["Date"].min().strftime("%Y-%m-%d"),
                    LastDate=df["Date"].max().strftime("%Y-%m-%d"),
                    Message="Downloaded successfully",
                )
                return df, status
        except Exception as e:
            last_msg = f"{type(e).__name__}: {e}"
            log(f"{ticker} {source} failed: {last_msg}")
            time.sleep(0.2)

    if cache_path.exists():
        try:
            df = pd.read_csv(cache_path, parse_dates=["Date"])
            df = normalize_ohlc(df, ticker, "cache")
            status.update(
                Status="STALE_OK",
                SourceUsed="cache",
                Rows=len(df),
                FirstDate=df["Date"].min().strftime("%Y-%m-%d"),
                LastDate=df["Date"].max().strftime("%Y-%m-%d"),
                Message=f"All online sources failed; using cache. Last error: {last_msg}",
            )
            return df, status
        except Exception as e:
            last_msg += f"; cache failed: {e}"

    status.update(Status="FAILED", SourceUsed="none", Message=last_msg)
    return None, status


def compute_period_summary(ohlc: pd.DataFrame, tickers_meta: pd.DataFrame, start: str, end: str, benchmark: str) -> pd.DataFrame:
    df = ohlc[(ohlc["Date"] >= pd.Timestamp(start)) & (ohlc["Date"] <= pd.Timestamp(end))].copy()
    rows = []
    bench_ret = None
    if benchmark in df["Ticker"].unique():
        b = df[df["Ticker"] == benchmark].sort_values("Date")
        if not b.empty:
            bench_ret = b["Close"].iloc[-1] / b["Close"].iloc[0] - 1

    for ticker, g in df.groupby("Ticker"):
        g = g.sort_values("Date").reset_index(drop=True)
        if g.empty:
            continue
        start_close = float(g["Close"].iloc[0])
        end_close = float(g["Close"].iloc[-1])
        close_return = end_close / start_close - 1 if start_close else None

        high_idx = g["High"].idxmax()
        low_idx = g["Low"].idxmin()
        period_high = float(g.loc[high_idx, "High"])
        period_low = float(g.loc[low_idx, "Low"])
        range_drawdown = period_low / period_high - 1 if period_high else None
        rebound_from_low = end_close / period_low - 1 if period_low else None
        current_vs_high = end_close / period_high - 1 if period_high else None
        half_recovery_level = period_low + 0.5 * (period_high - period_low)
        half_recovered = bool(end_close >= half_recovery_level)

        running_high = -math.inf
        running_peak_date = None
        worst_dd = 0.0
        worst_peak_date = None
        worst_trough_date = None
        for _, row in g.iterrows():
            if float(row["High"]) > running_high:
                running_high = float(row["High"])
                running_peak_date = row["Date"]
            if running_high > 0:
                dd = float(row["Low"]) / running_high - 1
                if dd < worst_dd:
                    worst_dd = dd
                    worst_peak_date = running_peak_date
                    worst_trough_date = row["Date"]

        meta = tickers_meta[tickers_meta["Ticker"] == ticker]
        name = meta["Name"].iloc[0] if not meta.empty else ""
        group = meta["Group"].iloc[0] if not meta.empty else ""
        role = meta["Role"].iloc[0] if not meta.empty else ""
        rel_vs_bench = close_return - bench_ret if (bench_ret is not None and close_return is not None) else None

        rows.append({
            "Ticker": ticker,
            "Name": name,
            "Group": group,
            "Role": role,
            "Start": start,
            "End": end,
            "Start_Close": start_close,
            "End_Close": end_close,
            "Close_Return_%": close_return * 100 if close_return is not None else None,
            "Relative_Return_vs_Benchmark_pp": rel_vs_bench * 100 if rel_vs_bench is not None else None,
            "Period_High": period_high,
            "Period_High_Date": g.loc[high_idx, "Date"].strftime("%Y-%m-%d"),
            "Period_Low": period_low,
            "Period_Low_Date": g.loc[low_idx, "Date"].strftime("%Y-%m-%d"),
            "High_to_Low_Range_%": range_drawdown * 100 if range_drawdown is not None else None,
            "Worst_Peak_to_Trough_Drawdown_%": worst_dd * 100,
            "Worst_DD_Peak_Date": "" if worst_peak_date is None else pd.Timestamp(worst_peak_date).strftime("%Y-%m-%d"),
            "Worst_DD_Trough_Date": "" if worst_trough_date is None else pd.Timestamp(worst_trough_date).strftime("%Y-%m-%d"),
            "Rebound_From_Period_Low_%": rebound_from_low * 100 if rebound_from_low is not None else None,
            "Current_vs_Period_High_%": current_vs_high * 100 if current_vs_high is not None else None,
            "Half_Recovery_Level": half_recovery_level,
            "Half_Recovered": half_recovered,
        })

    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values(["Worst_Peak_to_Trough_Drawdown_%", "Close_Return_%"], ascending=[False, False])
        out.insert(0, "Defense_Rank", range(1, len(out) + 1))
    return out


def json_records(df: pd.DataFrame) -> str:
    # DataFrame -> JSON records with ISO date strings
    x = df.copy()
    if "Date" in x.columns:
        x["Date"] = pd.to_datetime(x["Date"]).dt.strftime("%Y-%m-%d")
    return x.to_json(orient="records", force_ascii=False)


def build_dashboard_html(ohlc: pd.DataFrame, tickers_meta: pd.DataFrame, status: pd.DataFrame, config: dict) -> str:
    all_data_json = json_records(ohlc)
    meta_json = tickers_meta.to_json(orient="records", force_ascii=False)
    status_json = status.to_json(orient="records", force_ascii=False)
    benchmark = config.get("benchmark", "VT")
    min_date = pd.to_datetime(ohlc["Date"]).min().strftime("%Y-%m-%d")
    max_date = pd.to_datetime(ohlc["Date"]).max().strftime("%Y-%m-%d")
    default_months = int(config.get("dashboard", {}).get("default_months", 6))
    default_start = (pd.Timestamp(max_date) - pd.DateOffset(months=default_months)).strftime("%Y-%m-%d")
    default_all_visible = "true" if config.get("dashboard", {}).get("all_traces_visible_by_default", True) else "false"

    return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <title>ETF Leadership Dashboard</title>
  <script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script>
  <style>
    body {{ font-family: Segoe UI, Arial, sans-serif; margin: 22px; color: #222; }}
    h1 {{ margin-bottom: 4px; }}
    .subtle {{ color:#666; font-size: 13px; }}
    .controls {{ display:flex; flex-wrap:wrap; gap:10px; align-items:end; padding:12px; border:1px solid #ddd; border-radius:10px; background:#fafafa; }}
    .control-block {{ display:flex; flex-direction:column; gap:4px; }}
    label {{ font-size:12px; color:#444; }}
    input, select, button {{ padding:6px 8px; border:1px solid #bbb; border-radius:6px; background:white; }}
    button {{ cursor:pointer; }}
    .layout {{ display:grid; grid-template-columns: 300px 1fr; gap:16px; margin-top:14px; }}
    .panel {{ border:1px solid #ddd; border-radius:10px; padding:12px; max-height:760px; overflow:auto; }}
    #chart {{ width:100%; height:720px; }}
    .checkbox-row {{ display:flex; align-items:center; gap:6px; margin:4px 0; font-size:13px; }}
    .group-title {{ margin-top:10px; font-weight:700; border-top:1px solid #eee; padding-top:8px; }}
    table {{ border-collapse:collapse; width:100%; font-size:12px; }}
    th, td {{ border:1px solid #ddd; padding:5px 6px; text-align:right; }}
    th {{ background:#f2f2f2; position:sticky; top:0; z-index:1; }}
    td:first-child, th:first-child, td:nth-child(2), th:nth-child(2), td:nth-child(3), th:nth-child(3) {{ text-align:left; }}
    .section {{ margin-top:18px; }}
    .metric-note {{ padding:10px; background:#fff7ed; border:1px solid #fed7aa; border-radius:10px; margin:12px 0; font-size:13px; }}
    .good {{ color:#047857; font-weight:600; }}
    .bad {{ color:#B91C1C; font-weight:600; }}
    .yes {{ color:#047857; font-weight:700; }}
    .no {{ color:#B91C1C; font-weight:700; }}
  </style>
</head>
<body>
  <h1>ETF Leadership Dashboard</h1>
  <div class="subtle">Generated at {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} · Data range: {min_date} to {max_date}</div>
  <div class="metric-note">
    <b>중요:</b> 기간 내 최고점은 <b>daily High의 최대값</b>, 최저점은 <b>daily Low의 최소값</b>으로 계산한다.
    Worst drawdown은 매일의 Low를 그 이전까지의 running High와 비교한다. 따라서 종가 기준이 아니다.
    단, 라인 차트의 기본 모드는 가독성을 위해 normalized Close를 사용한다.
  </div>

  <div class="controls">
    <div class="control-block">
      <label>Start date</label>
      <input type="date" id="startDate" min="{min_date}" max="{max_date}" value="{default_start}">
    </div>
    <div class="control-block">
      <label>End date</label>
      <input type="date" id="endDate" min="{min_date}" max="{max_date}" value="{max_date}">
    </div>
    <div class="control-block">
      <label>Benchmark</label>
      <select id="benchmarkSelect"></select>
    </div>
    <div class="control-block">
      <label>Chart mode</label>
      <select id="chartMode">
        <option value="normalized_close">Normalized Close, start=100</option>
        <option value="drawdown_low">Intraday drawdown from running High, %</option>
        <option value="relative_strength">Relative strength vs benchmark, start=100</option>
      </select>
    </div>
    <div class="control-block">
      <label>Table sort metric</label>
      <select id="sortMetric">
        <option value="Worst_Drawdown">Worst DD</option>
        <option value="High_to_Low_Range">High→Low Range</option>
        <option value="Rebound_From_Low">Rebound from Low</option>
        <option value="Current_vs_High">Current vs High</option>
        <option value="Close_Return">Close Return</option>
        <option value="Relative_Return">Vs Bench</option>
      </select>
    </div>
    <div class="control-block">
      <label>Sort direction</label>
      <select id="sortDirection">
        <option value="desc">내림차순: 큰 값 먼저</option>
        <option value="asc">오름차순: 작은 값 먼저</option>
      </select>
    </div>
    <button onclick="updateDashboard()">Update</button>
    <button onclick="setLastMonths(1)">1M</button>
    <button onclick="setLastMonths(3)">3M</button>
    <button onclick="setLastMonths(6)">6M</button>
    <button onclick="setLastMonths(12)">1Y</button>
    <button onclick="setFullRange()">Full</button>
  </div>

  <div class="layout">
    <div class="panel">
      <div style="display:flex; flex-wrap:wrap; gap:6px;">
        <button onclick="setVisibilityAll(true)">Show all</button>
        <button onclick="setVisibilityAll(false)">Hide all</button>
        <button onclick="showGroup('US Sector')">US sectors</button>
        <button onclick="showGroup('Industry')">Industries</button>
        <button onclick="showGroup('Global Region')">Regions</button>
        <button onclick="showRoleContains('Defensive')">Defensive</button>
        <button onclick="showRoleContains('Cyclical')">Cyclicals</button>
      </div>
      <div id="tickerChecks"></div>
    </div>
    <div>
      <div id="chart"></div>
    </div>
  </div>

  <div class="section">
    <h2>Period summary</h2>
    <div class="subtle">High/Low and worst drawdown use daily High/Low. Click table headers are not sortable; use Table sort selector above.</div>
    <div id="summaryTable"></div>
  </div>

  <div class="section">
    <h2>Download status</h2>
    <div id="statusTable"></div>
  </div>

<script>
const RAW = {all_data_json};
const META = {meta_json};
const STATUS = {status_json};
const DEFAULT_BENCHMARK = "{benchmark}";
const DEFAULT_ALL_VISIBLE = {default_all_visible};

const byTicker = {{}};
for (const r of RAW) {{
  if (!byTicker[r.Ticker]) byTicker[r.Ticker] = [];
  byTicker[r.Ticker].push(r);
}}
for (const t in byTicker) {{
  byTicker[t].sort((a,b) => a.Date.localeCompare(b.Date));
}}

const visible = {{}};
for (const m of META) visible[m.Ticker] = DEFAULT_ALL_VISIBLE;

function fmtPct(x) {{
  if (x === null || x === undefined || Number.isNaN(x)) return "";
  const cls = x >= 0 ? "good" : "bad";
  return `<span class="${{cls}}">${{x.toFixed(2)}}%</span>`;
}}
function fmtNum(x) {{
  if (x === null || x === undefined || Number.isNaN(x)) return "";
  return Number(x).toFixed(2);
}}
function filterPeriod(rows, start, end) {{
  return rows.filter(r => r.Date >= start && r.Date <= end);
}}
function firstClose(rows) {{
  return rows.length ? Number(rows[0].Close) : null;
}}
function computeMetrics(ticker, rows, benchmarkReturn) {{
  if (!rows.length) return null;
  const startClose = Number(rows[0].Close);
  const endClose = Number(rows[rows.length-1].Close);
  const closeReturn = startClose ? (endClose/startClose - 1) * 100 : null;

  let periodHigh = -Infinity, highDate = "";
  let periodLow = Infinity, lowDate = "";
  for (const r of rows) {{
    const h = Number(r.High), l = Number(r.Low);
    if (h > periodHigh) {{ periodHigh = h; highDate = r.Date; }}
    if (l < periodLow) {{ periodLow = l; lowDate = r.Date; }}
  }}
  const highLowRange = periodHigh ? (periodLow/periodHigh - 1) * 100 : null;
  const reboundFromLow = periodLow ? (endClose/periodLow - 1) * 100 : null;
  const currentVsHigh = periodHigh ? (endClose/periodHigh - 1) * 100 : null;
  const halfRecoveryLevel = periodLow + 0.5 * (periodHigh - periodLow);
  const halfRecovered = endClose >= halfRecoveryLevel;

  let runningHigh = -Infinity, runningPeakDate = "";
  let worstDD = 0, worstPeakDate = "", worstTroughDate = "";
  for (const r of rows) {{
    const h = Number(r.High), l = Number(r.Low);
    if (h > runningHigh) {{ runningHigh = h; runningPeakDate = r.Date; }}
    const dd = runningHigh ? (l/runningHigh - 1) * 100 : 0;
    if (dd < worstDD) {{
      worstDD = dd; worstPeakDate = runningPeakDate; worstTroughDate = r.Date;
    }}
  }}

  const meta = META.find(x => x.Ticker === ticker) || {{}};
  const relative = benchmarkReturn === null ? null : closeReturn - benchmarkReturn;

  return {{
    Ticker: ticker,
    Name: meta.Name || "",
    Group: meta.Group || "",
    Role: meta.Role || "",
    Start_Close: startClose,
    End_Close: endClose,
    Close_Return: closeReturn,
    Relative_Return: relative,
    Period_High: periodHigh,
    Period_High_Date: highDate,
    Period_Low: periodLow,
    Period_Low_Date: lowDate,
    High_to_Low_Range: highLowRange,
    Worst_Drawdown: worstDD,
    Worst_DD_Peak_Date: worstPeakDate,
    Worst_DD_Trough_Date: worstTroughDate,
    Rebound_From_Low: reboundFromLow,
    Current_vs_High: currentVsHigh,
    Half_Recovery_Level: halfRecoveryLevel,
    Half_Recovered: halfRecovered
  }};
}}

function computeSeries(ticker, rows, mode, benchRows) {{
  if (!rows.length) return {{x:[], y:[]}};
  const x = rows.map(r => r.Date);
  let y = [];
  if (mode === "normalized_close") {{
    const base = Number(rows[0].Close);
    y = rows.map(r => Number(r.Close)/base*100);
  }} else if (mode === "drawdown_low") {{
    let rh = -Infinity;
    y = rows.map(r => {{
      const h = Number(r.High), l = Number(r.Low);
      if (h > rh) rh = h;
      return rh ? (l/rh - 1)*100 : 0;
    }});
  }} else if (mode === "relative_strength") {{
    if (!benchRows || !benchRows.length) {{
      const base = Number(rows[0].Close);
      y = rows.map(r => Number(r.Close)/base*100);
    }} else {{
      const bmap = {{}};
      for (const b of benchRows) bmap[b.Date] = Number(b.Close);
      const tickerBase = Number(rows[0].Close);
      const benchBase = Number(benchRows[0].Close);
      y = rows.map(r => {{
        const b = bmap[r.Date];
        if (!b || !tickerBase || !benchBase) return null;
        return (Number(r.Close)/tickerBase)/(b/benchBase)*100;
      }});
    }}
  }}
  return {{x, y}};
}}

function populateControls() {{
  const bench = document.getElementById("benchmarkSelect");
  for (const m of META) {{
    const opt = document.createElement("option");
    opt.value = m.Ticker; opt.textContent = `${{m.Ticker}} - ${{m.Name}}`;
    if (m.Ticker === DEFAULT_BENCHMARK) opt.selected = true;
    bench.appendChild(opt);
  }}

  const box = document.getElementById("tickerChecks");
  const groups = [...new Set(META.map(m => m.Group))];
  for (const g of groups) {{
    const title = document.createElement("div");
    title.className = "group-title";
    title.textContent = g;
    box.appendChild(title);
    for (const m of META.filter(x => x.Group === g)) {{
      const row = document.createElement("div");
      row.className = "checkbox-row";
      const cb = document.createElement("input");
      cb.type = "checkbox"; cb.id = "cb_" + m.Ticker; cb.checked = visible[m.Ticker];
      cb.onchange = () => {{ visible[m.Ticker] = cb.checked; updateDashboard(); }};
      const lab = document.createElement("label");
      lab.htmlFor = cb.id;
      lab.textContent = `${{m.Ticker}} · ${{m.Role}}`;
      row.appendChild(cb); row.appendChild(lab); box.appendChild(row);
    }}
  }}
}}

function setVisibilityAll(v) {{
  for (const t in visible) visible[t] = v;
  for (const t in visible) {{
    const cb = document.getElementById("cb_" + t);
    if (cb) cb.checked = v;
  }}
  updateDashboard();
}}
function showGroup(groupName) {{
  for (const m of META) visible[m.Ticker] = (m.Group === groupName);
  syncChecks(); updateDashboard();
}}
function showRoleContains(text) {{
  for (const m of META) visible[m.Ticker] = (m.Role || "").includes(text);
  syncChecks(); updateDashboard();
}}
function syncChecks() {{
  for (const t in visible) {{
    const cb = document.getElementById("cb_" + t);
    if (cb) cb.checked = visible[t];
  }}
}}
function setLastMonths(n) {{
  const end = document.getElementById("endDate").value;
  const d = new Date(end);
  d.setMonth(d.getMonth() - n);
  document.getElementById("startDate").value = d.toISOString().slice(0,10);
  updateDashboard();
}}
function setFullRange() {{
  document.getElementById("startDate").value = "{min_date}";
  document.getElementById("endDate").value = "{max_date}";
  updateDashboard();
}}

function renderSummary(metrics) {{
  const sortMetric = document.getElementById("sortMetric").value;
  const sortDirection = document.getElementById("sortDirection").value;
  const arr = [...metrics];

  arr.sort((a,b) => {{
    const av = a[sortMetric];
    const bv = b[sortMetric];

    const aMissing = (av === null || av === undefined || Number.isNaN(av));
    const bMissing = (bv === null || bv === undefined || Number.isNaN(bv));
    if (aMissing && bMissing) return 0;
    if (aMissing) return 1;
    if (bMissing) return -1;

    return sortDirection === "asc" ? av - bv : bv - av;
  }});

  let html = "<table><thead><tr>" +
    "<th>Rank</th><th>Ticker</th><th>Group</th><th>Role</th>" +
    "<th>Close Return</th><th>Vs Bench</th><th>Worst DD<br>High→Low</th>" +
    "<th>DD Peak</th><th>DD Trough</th><th>Period High</th><th>High Date</th>" +
    "<th>Period Low</th><th>Low Date</th><th>High→Low Range</th>" +
    "<th>Rebound from Low</th><th>Current vs High</th><th>50% Recovered?</th>" +
    "</tr></thead><tbody>";
  arr.forEach((m, idx) => {{
    html += `<tr>
      <td>${{idx+1}}</td>
      <td><b>${{m.Ticker}}</b></td>
      <td>${{m.Group}}</td>
      <td>${{m.Role}}</td>
      <td>${{fmtPct(m.Close_Return)}}</td>
      <td>${{fmtPct(m.Relative_Return)}}</td>
      <td>${{fmtPct(m.Worst_Drawdown)}}</td>
      <td>${{m.Worst_DD_Peak_Date}}</td>
      <td>${{m.Worst_DD_Trough_Date}}</td>
      <td>${{fmtNum(m.Period_High)}}</td>
      <td>${{m.Period_High_Date}}</td>
      <td>${{fmtNum(m.Period_Low)}}</td>
      <td>${{m.Period_Low_Date}}</td>
      <td>${{fmtPct(m.High_to_Low_Range)}}</td>
      <td>${{fmtPct(m.Rebound_From_Low)}}</td>
      <td>${{fmtPct(m.Current_vs_High)}}</td>
      <td>${{m.Half_Recovered ? '<span class="yes">Yes</span>' : '<span class="no">No</span>'}}</td>
    </tr>`;
  }});
  html += "</tbody></table>";
  document.getElementById("summaryTable").innerHTML = html;
}}

function renderStatus() {{
  let html = "<table><thead><tr><th>Ticker</th><th>Status</th><th>Source</th><th>Rows</th><th>First</th><th>Last</th><th>Message</th></tr></thead><tbody>";
  for (const s of STATUS) {{
    html += `<tr><td>${{s.Ticker}}</td><td>${{s.Status}}</td><td>${{s.SourceUsed}}</td><td>${{s.Rows}}</td><td>${{s.FirstDate}}</td><td>${{s.LastDate}}</td><td>${{s.Message}}</td></tr>`;
  }}
  html += "</tbody></table>";
  document.getElementById("statusTable").innerHTML = html;
}}

function updateDashboard() {{
  const start = document.getElementById("startDate").value;
  const end = document.getElementById("endDate").value;
  const benchmark = document.getElementById("benchmarkSelect").value;
  const mode = document.getElementById("chartMode").value;

  const benchRows = filterPeriod(byTicker[benchmark] || [], start, end);
  let benchReturn = null;
  if (benchRows.length) {{
    benchReturn = (Number(benchRows[benchRows.length-1].Close)/Number(benchRows[0].Close)-1)*100;
  }}

  const traces = [];
  const metrics = [];
  for (const m of META) {{
    const rows = filterPeriod(byTicker[m.Ticker] || [], start, end);
    if (!rows.length) continue;
    const metric = computeMetrics(m.Ticker, rows, benchReturn);
    if (metric) metrics.push(metric);
    const series = computeSeries(m.Ticker, rows, mode, benchRows);
    traces.push({{
      x: series.x, y: series.y, type: "scatter", mode: "lines",
      name: m.Ticker,
      visible: visible[m.Ticker] ? true : "legendonly",
      hovertemplate: m.Ticker + "<br>%{{x}}<br>%{{y:.2f}}<extra></extra>"
    }});
  }}

  const ytitle = mode === "normalized_close" ? "Normalized Close, start=100"
    : mode === "drawdown_low" ? "Drawdown from running High using daily Low, %"
    : "Relative Strength vs " + benchmark + ", start=100";

  Plotly.newPlot("chart", traces, {{
    title: `ETF movement overlay · ${{start}} to ${{end}} · benchmark ${{benchmark}}`,
    xaxis: {{title: "Date", rangeslider: {{visible: true}}}},
    yaxis: {{title: ytitle, zeroline: true}},
    legend: {{orientation: "v", x: 1.02, y: 1}},
    margin: {{l: 60, r: 160, t: 50, b: 50}},
    hovermode: "x unified"
  }}, {{responsive: true}});

  renderSummary(metrics);
  renderStatus();
}}

populateControls();
document.getElementById("sortMetric").addEventListener("change", updateDashboard);
document.getElementById("sortDirection").addEventListener("change", updateDashboard);
updateDashboard();
</script>
</body>
</html>"""


def write_excel(ohlc: pd.DataFrame, meta: pd.DataFrame, status: pd.DataFrame, summary: pd.DataFrame) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.chart import LineChart, Reference
    except Exception as e:
        log(f"openpyxl unavailable; skipping Excel: {e}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws.append(["ETF Leadership Dashboard"])
    ws.append(["Use output/dashboard.html for interactive toggle/date controls."])
    ws.append(["Period high = max daily High. Period low = min daily Low. Worst drawdown = daily Low / prior running High - 1."])
    ws.append(["This workbook is generated by GitHub Actions."])
    ws["A1"].font = Font(bold=True, size=16)

    ws2 = wb.create_sheet("Default_Period_Summary")
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws2.append(r)

    ws3 = wb.create_sheet("Tickers")
    for r in dataframe_to_rows(meta, index=False, header=True):
        ws3.append(r)

    ws4 = wb.create_sheet("Download_Status")
    for r in dataframe_to_rows(status, index=False, header=True):
        ws4.append(r)

    ws5 = wb.create_sheet("Prices_OHLC")
    # Keep full long data; 80k rows is okay for these tickers/timeframe.
    out = ohlc.copy()
    out["Date"] = pd.to_datetime(out["Date"]).dt.strftime("%Y-%m-%d")
    for r in dataframe_to_rows(out, index=False, header=True):
        ws5.append(r)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet in [ws2, ws3, ws4, ws5]:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 30)
        sheet.freeze_panes = "A2"

    # Add chart on summary sheet for worst drawdown and close return if columns exist
    try:
        # Find columns
        headers = [c.value for c in ws2[1]]
        if "Ticker" in headers and "Worst_Peak_to_Trough_Drawdown_%" in headers and "Close_Return_%" in headers:
            ticker_col = headers.index("Ticker") + 1
            dd_col = headers.index("Worst_Peak_to_Trough_Drawdown_%") + 1
            ret_col = headers.index("Close_Return_%") + 1
            max_row = min(ws2.max_row, 40)
            chart = LineChart()
            chart.title = "Default period: Return vs Worst DD"
            chart.y_axis.title = "%"
            chart.x_axis.title = "Ticker rank"
            data1 = Reference(ws2, min_col=dd_col, min_row=1, max_row=max_row)
            data2 = Reference(ws2, min_col=ret_col, min_row=1, max_row=max_row)
            chart.add_data(data1, titles_from_data=True)
            chart.add_data(data2, titles_from_data=True)
            cats = Reference(ws2, min_col=ticker_col, min_row=2, max_row=max_row)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 18
            ws2.add_chart(chart, "W2")
    except Exception:
        pass

    wb.save(ROOT / "etf_leadership_dashboard.xlsx")
    log("Created etf_leadership_dashboard.xlsx")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default=str(ROOT / "config.json"))
    args = parser.parse_args()

    LOG_PATH.write_text("", encoding="utf-8")

    config = load_config(Path(args.config))
    start_date = config.get("start_date", "2018-01-01")
    end_date = pd.Timestamp.today().strftime("%Y-%m-%d")
    source_order = config.get("data_source_order", ["stooq", "yahoo"])

    meta_rows = []
    all_frames = []
    statuses = []
    for item in config["tickers"]:
        ticker = item["ticker"].upper()
        meta_rows.append({
            "Ticker": ticker,
            "Name": item.get("name", ""),
            "Group": item.get("group", ""),
            "Role": item.get("role", ""),
        })
        df, st = get_ohlc_for_ticker(ticker, start_date, end_date, source_order)
        statuses.append(st)
        if df is not None and not df.empty:
            all_frames.append(df)

    meta = pd.DataFrame(meta_rows)
    status = pd.DataFrame(statuses)

    if not all_frames:
        raise RuntimeError("No OHLC data downloaded for any ticker.")

    ohlc = pd.concat(all_frames, ignore_index=True)
    ohlc = ohlc.merge(meta, on="Ticker", how="left")
    ohlc = ohlc.sort_values(["Ticker", "Date"])

    # Export long data
    long_out = ohlc.copy()
    long_out["Date"] = pd.to_datetime(long_out["Date"]).dt.strftime("%Y-%m-%d")
    long_out.to_csv(OUTPUT / "prices_ohlc.csv", index=False, encoding="utf-8-sig")
    status.to_csv(OUTPUT / "download_status.csv", index=False, encoding="utf-8-sig")
    meta.to_csv(OUTPUT / "tickers.csv", index=False, encoding="utf-8-sig")

    max_date = pd.to_datetime(ohlc["Date"]).max()
    default_start = (max_date - pd.DateOffset(months=int(config.get("dashboard", {}).get("default_months", 6)))).strftime("%Y-%m-%d")
    default_end = max_date.strftime("%Y-%m-%d")
    summary = compute_period_summary(ohlc, meta, default_start, default_end, config.get("benchmark", "VT"))
    summary.to_csv(OUTPUT / "default_period_summary.csv", index=False, encoding="utf-8-sig")

    html = build_dashboard_html(ohlc[["Date","Ticker","Open","High","Low","Close","Volume","Source"]], meta, status, config)
    (OUTPUT / "dashboard.html").write_text(html, encoding="utf-8")

    # Simple machine-readable manifest
    manifest = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "start_date": start_date,
        "end_date": end_date,
        "tickers": [x["ticker"] for x in config["tickers"]],
        "benchmark": config.get("benchmark", "VT"),
        "outputs": [
            "output/prices_ohlc.csv",
            "output/default_period_summary.csv",
            "output/download_status.csv",
            "output/dashboard.html",
            "etf_leadership_dashboard.xlsx",
        ],
        "high_low_rule": config.get("notes", {}).get("high_low_rule", ""),
    }
    (OUTPUT / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")

    write_excel(ohlc, meta, status, summary)
    log("Done. Check output/dashboard.html and etf_leadership_dashboard.xlsx")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
